from __future__ import annotations

import os
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill # <--- Added Import
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# --- Configuration ---
MODULE_DIR = Path(__file__).resolve().parent
CASES_PATH = MODULE_DIR / "auth_test_cases.xlsx"
REPORT_DIR = MODULE_DIR / "reports"
REPORT_FILE = REPORT_DIR / "auth_results.xlsx"
REPORT_HEADERS = [
    "TimestampUTC", "CaseID", "Feature", "Scenario", "Status", "Notes",
]

load_dotenv()

# --- Helpers: Report Management ---

def _ensure_report_directory() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def _load_auth_cases() -> List[Dict[str, str]]:
    print(f"\n[DEBUG] Finding test case file at: {CASES_PATH}")
    print(f"[DEBUG] File exists? {CASES_PATH.exists()}\n")
    if not CASES_PATH.exists():
        pytest.skip(f"Auth cases file not found at {CASES_PATH}")

    workbook = load_workbook(CASES_PATH)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    cases: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        entry: Dict[str, str] = {header: value for header, value in zip(headers, row)}
        cases.append(_normalize_case_entry(entry))

    if len(cases) < 1:
        pytest.skip("Auth cases workbook does not contain any rows to execute")

    return cases


def _append_report_row(case: Dict[str, str], status: str, notes: str) -> None:
    _ensure_report_directory()

    if REPORT_FILE.exists():
        workbook = load_workbook(REPORT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(REPORT_HEADERS)
        # Format bold header for new file
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    # Current row data
    row_data = [
        datetime.now(UTC).isoformat(timespec="seconds"),
        case.get("CaseID"),
        case.get("Feature"),
        case.get("Scenario"),
        status,
        notes,
    ]

    sheet.append(row_data)

    # --- LOGIC FORMAT RED IF FAIL ---
    if status == "FAIL":
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        
        # Get the last added row index
        last_row_index = sheet.max_row
        
        # Color each cell in that row
        for col_idx in range(1, len(row_data) + 1):
            cell = sheet.cell(row=last_row_index, column=col_idx)
            cell.fill = red_fill
            cell.font = red_font

    # Auto-adjust column width (for readability)
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(REPORT_FILE)


def _normalize_case_entry(entry: Dict[str, str]) -> Dict[str, str]:
    normalized: Dict[str, str] = {}
    for key, value in entry.items():
        if value is None:
            normalized[key] = ""
        elif isinstance(value, str):
            normalized[key] = value.strip()
        else:
            normalized[key] = str(value)
    return normalized


def _resolve_dynamic_value(value: str) -> str:
    if not value:
        return ""
    resolved = value
    if "{{ts}}" in resolved:
        resolved = resolved.replace("{{ts}}", datetime.now(UTC).strftime("%Y%m%d%H%M%S"))
    return resolved

# --- Helpers: Selenium & State Management ---

def _logout_if_possible(driver) -> None:
    try:
        wait = WebDriverWait(driver, 5)
        avatar = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "img[alt='User']"))
        )
        avatar.click()
        logout_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Logout')]") )
        )
        logout_button.click()
        wait.until(EC.visibility_of_element_located((By.XPATH, "//h2[text()='Login']")))
    except TimeoutException:
        pass
    except Exception:
        pass


def _get_wait(driver, timeout: int = 8) -> WebDriverWait:
    return WebDriverWait(driver, timeout)


def _clear_existing_report() -> None:
    _ensure_report_directory()
    if REPORT_FILE.exists():
        REPORT_FILE.unlink()


# --- Pytest Fixtures ---

@pytest.fixture(scope="session")
def base_url() -> str:
    url = os.getenv("E2E_BASE_URL") or "https://cinestech.me"
    return url.rstrip("/")


@pytest.fixture(scope="session")
def driver():
    options = Options()

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1440,900")
    
    # Optional: Add headless mode if needed
    # options.add_argument("--headless")

    service = Service(ChromeDriverManager().install())
    browser = webdriver.Chrome(service=service, options=options)

    yield browser

    browser.quit()


@pytest.fixture(scope="session", autouse=True)
def reset_auth_report() -> None:
    _clear_existing_report()


@pytest.fixture(scope="session")
def auth_case_data() -> List[Dict[str, str]]:
    return _load_auth_cases()


# --- Core Test Execution Logic ---

def _execute_auth_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    feature = case.get("Feature", "").lower()

    if feature == "login":
        return _run_login_case(driver, base_url, case)
    if feature == "forgot password":
        return _run_forgot_case(driver, base_url, case)
    
    # REGISTER LOGIC REMOVED
    
    return "SKIP", f"Unsupported feature: {case.get('Feature')}"


def _fill_if_present(element, value: str) -> None:
    if value:
        element.clear()
        element.send_keys(value)
    else:
        element.clear()


def _get_assert_type(case: Dict[str, str]) -> str:
    return (case.get("AssertType") or "").strip().lower()


def _run_login_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    wait = _get_wait(driver)
    login_url = f"{base_url}/login"
    driver.get(login_url)

    email_input = wait.until(EC.visibility_of_element_located((By.ID, "email")))
    password_input = driver.find_element(By.ID, "password")
    submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")

    assert_type = _get_assert_type(case)

    email_value = _resolve_dynamic_value(case.get("Data_Email", ""))
    password_value = case.get("Data_Password", "")

    # Logic to handle 'html5_required' for both fields
    if assert_type == "html5_required" and not email_value and not password_value:
        email_input.clear()
        password_input.clear()
    elif assert_type == "html5_required" and email_value and not password_value:
         _fill_if_present(email_input, email_value)
         password_input.clear()
    else:
         _fill_if_present(email_input, email_value)
         _fill_if_present(password_input, password_value)


    remember_value = (case.get("Data_RememberMe") or "").strip().upper()
    if remember_value == "TRUE":
        remember_checkbox = driver.find_element(By.ID, "remember-me")
        if not remember_checkbox.is_selected():
            remember_checkbox.click()

    submit_button.click()
    time.sleep(0.5)

    if assert_type == "html5_required":
        target_input = password_input if email_value else email_input
        validation_message = driver.execute_script("return arguments[0].validationMessage;", target_input)
        if validation_message:
            return "PASS", f"HTML required validation triggered: {validation_message}"
        return "FAIL", "Expected HTML5 required validation message but none was found"

    if assert_type == "html5_type_mismatch":
        validation_message = driver.execute_script("return arguments[0].validationMessage;", email_input)
        if validation_message:
            return "PASS", f"HTML email validation triggered: {validation_message}"
        return "FAIL", "Expected email format validation message but got none"

    if assert_type == "error_banner":
        expected_message = (case.get("ExpectedUIMessage") or "").lower()
        try:
            error_box = wait.until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "//div[contains(@class,'text-red') or contains(@class,'bg-red')]"
                    )
                )
            )
            error_text = error_box.text.strip()
            if expected_message and expected_message not in error_text.lower():
                return (
                    "FAIL",
                    f"Error text mismatch. Expected fragment '{expected_message}', got '{error_text}'",
                )
            return "PASS", f"Error banner displayed: {error_text}"
        except TimeoutException:
            return "FAIL", "Expected error banner did not appear"

    # Success-oriented flows
    try:
        wait.until(lambda d: "/login" not in d.current_url)
    except TimeoutException:
        return "FAIL", "Did not navigate away from login page"

    current_url = driver.current_url

    if assert_type == "remember_me":
        # Treat remember me as successful login with checkbox interaction verified
        notes = f"Login succeeded with remember me enabled; redirected to {current_url}"
        _logout_if_possible(driver)
        return "PASS", notes

    if assert_type == "login_success":
        if "email-verification" in current_url:
            return "PASS", "Redirected to email verification flow"

        try:
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "img[alt='User']")))
            notes = f"User avatar visible after login; current url {current_url}"
            _logout_if_possible(driver)
            return "PASS", notes
        except TimeoutException:
            return "FAIL", f"Login expected to succeed but user avatar not found (url={current_url})"

    return "FAIL", f"Unhandled assertion type for login: {case.get('AssertType')}"


def _run_forgot_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    wait = _get_wait(driver)
    forgot_url = f"{base_url}/forgot-password"
    driver.get(forgot_url)

    email_input = wait.until(EC.visibility_of_element_located((By.ID, "email")))
    reset_email = _resolve_dynamic_value(case.get("Data_ResetEmail", ""))
    
    assert_type = _get_assert_type(case)

    if assert_type != "html5_required":
        _fill_if_present(email_input, reset_email)
    else:
        email_input.clear()

    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    time.sleep(0.5)

    expected_message = (case.get("ExpectedUIMessage") or "").lower()

    if assert_type == "html5_required" or assert_type == "html5_type_mismatch":
        validation_message = driver.execute_script("return arguments[0].validationMessage;", email_input)
        if validation_message:
            return "PASS", f"HTML validation triggered: {validation_message}"
        return "FAIL", "Expected HTML5 validation message but none was found"

    if assert_type == "forgot_success":
        try:
            wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//h2[contains(., 'Check Your Email')]")
                )
            )
            notes = "Forgot password success card displayed"
            try:
                driver.find_element(By.XPATH, "//button[contains(., 'Back to Login')]").click()
            except Exception:
                pass
            return "PASS", notes
        except TimeoutException:
            return "FAIL", "Success message for forgot password did not appear"

    if assert_type == "error_banner":
        try:
            error_box = wait.until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "//div[contains(@class,'text-red') or contains(@class,'bg-red')]"
                    )
                )
            )
            error_text = error_box.text.strip()
            if expected_message and expected_message not in error_text.lower():
                return (
                    "FAIL",
                    f"Forgot password error mismatch. Expected '{expected_message}', got '{error_text}'",
                )
            return "PASS", f"Forgot password error displayed: {error_text}"
        except TimeoutException:
            return "FAIL", "Expected forgot-password error banner did not appear"

    return "SKIP", f"Unhandled assert type for forgot password: {case.get('AssertType')}"


# --- REGISTER FUNCTION REMOVED ---


# --- Main Test Runner ---

AUTH_CASES = _load_auth_cases()


@pytest.mark.parametrize(
    "case_entry",
    AUTH_CASES,
    ids=lambda case: case.get("CaseID", "unknown"),
)
@pytest.mark.e2e
def test_auth_flow(case_entry: Dict[str, str], driver, base_url) -> None:
    case = _normalize_case_entry(case_entry)
    status = "ERROR"
    notes = ""

    try:
        status, notes = _execute_auth_case(driver, base_url, case)
    except Exception as exc:  # noqa: BLE001 - capture any failure for reporting
        status = "FAIL"
        notes = f"{type(exc).__name__}: {exc}"
    finally:
        _append_report_row(case, status, notes)

    if status == "PASS":
        return
    if status == "SKIP":
        pytest.skip(notes or "Scenario skipped")

    pytest.fail(notes or "Scenario failed")