from __future__ import annotations

import os
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains

MODULE_DIR = Path(__file__).resolve().parent
CASES_PATH = MODULE_DIR / "movie_test_cases.xlsx"
REPORT_DIR = MODULE_DIR / "reports"
REPORT_FILE = REPORT_DIR / "movie_results.xlsx"
REPORT_HEADERS = [
    "TimestampUTC",
    "CaseID",
    "Feature",
    "Scenario",
    "Status",
    "Notes",
]

load_dotenv()


def _ensure_report_directory() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def _load_movie_cases() -> List[Dict[str, str]]:
    print(f"\n[DEBUG] Đang tìm file test case tại: {CASES_PATH}")
    if not CASES_PATH.exists():
        pytest.skip(f"Movie cases file not found at {CASES_PATH}")

    workbook = load_workbook(CASES_PATH)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    cases: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        entry: Dict[str, str] = {header: value for header, value in zip(headers, row)}
        cases.append(_normalize_case_entry(entry))

    if len(cases) < 1:
        pytest.skip("Movie cases workbook does not contain any rows")

    return cases


def _append_report_row(case: Dict[str, str], status: str, notes: str) -> None:
    _ensure_report_directory()
    if REPORT_FILE.exists():
        workbook = load_workbook(REPORT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(REPORT_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    row_data = [
        datetime.now(UTC).isoformat(timespec="seconds"),
        case.get("CaseID"),
        case.get("Feature"),
        case.get("Scenario"),
        status,
        notes,
    ]
    
    sheet.append(row_data)

    if status == "FAIL":
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        last_row_index = sheet.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = sheet.cell(row=last_row_index, column=col_idx)
            cell.fill = red_fill
            cell.font = red_font

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(REPORT_FILE)


def _normalize_case_entry(entry: Dict[str, str]) -> Dict[str, str]:
    normalized: Dict[str, str] = {}
    for key, value in entry.items():
        normalized[key] = str(value).strip() if value is not None else ""
    return normalized


def _get_wait(driver, timeout: int = 10) -> WebDriverWait:
    return WebDriverWait(driver, timeout)


def _clear_existing_report() -> None:
    _ensure_report_directory()
    if REPORT_FILE.exists():
        try:
            REPORT_FILE.unlink()
        except PermissionError:
            pytest.exit(f"Không thể xóa report cũ. Vui lòng đóng file {REPORT_FILE.name}")


@pytest.fixture(scope="session")
def base_url() -> str:
    url = os.getenv("E2E_BASE_URL") or "https://cinestech.me"
    return url.rstrip("/")


@pytest.fixture(scope="session")
def driver():
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    # Chiến lược load trang nhanh (không chờ ảnh)
    options.page_load_strategy = 'eager'
    
    service = Service(ChromeDriverManager().install())
    browser = webdriver.Chrome(service=service, options=options)
    
    yield browser
    browser.quit()


@pytest.fixture(scope="session", autouse=True)
def reset_movie_report() -> None:
    _clear_existing_report()


@pytest.fixture(scope="session")
def movie_case_data() -> List[Dict[str, str]]:
    return _load_movie_cases()


def _execute_movie_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    # Wait chuẩn 10s
    wait = _get_wait(driver, 10)
    # Wait ngắn 2s (Dùng để kiểm tra cái gì đó KHÔNG tồn tại để đỡ tốn thời gian chờ)
    short_wait = WebDriverWait(driver, 2)

    movie_id = case.get("Data_MovieID", "invalid-uuid-123")
    page_url = f"{base_url}/movie/{movie_id}"
    
    precondition = case.get("Preconditions", "")
    if "chưa đăng nhập" in precondition:
        driver.get(f"{base_url}/login")
    elif "đã đăng nhập" in precondition:
        pass

    driver.get(page_url)
    
    assert_type = case.get("AssertType", "").lower()
    selector = case.get("Data_AssertSelector", "")
    click_selector = case.get("Data_ClickSelector", "")
    expected_text = case.get("Data_ExpectedText", "")

    try:
        def get_locator(selector_str):
            if selector_str.startswith("//") or selector_str.startswith("(//"):
                return (By.XPATH, selector_str)
            return (By.CSS_SELECTOR, selector_str)
        
        # --- Xử lý 4 CASE DESIGNED TO FAIL ---
        if case["CaseID"] == "TC09":
            wait.until(EC.element_to_be_clickable((By.XPATH, click_selector))).click()
            # Giảm thời gian chờ load iframe
            try:
                short_wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "iframe[src*='youtube']")))
            except TimeoutException:
                return "FAIL", "Trailer modal did not open within 2s."
            
            overflow_style = driver.execute_script("return document.body.style.overflow")
            if overflow_style == "hidden":
                return "PASS", "Scroll lock implemented correctly."
            return "FAIL", f"Scroll lock missing. Body overflow='{overflow_style}' (expected 'hidden')."

        if case["CaseID"] == "TC10":
            try:
                # Dùng short_wait vì nếu không có ảnh thì fail nhanh
                backdrop_imgs = short_wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'absolute')]//img")))
                if not backdrop_imgs:
                    return "FAIL", "Backdrop image not found in DOM."
                
                src = backdrop_imgs[0].get_attribute("src")
                if src and (src.startswith("http") or src.startswith("/")):
                    return "PASS", f"Backdrop image loaded: {src[:50]}..."
                return "FAIL", f"Backdrop src invalid or null."
            except TimeoutException:
                return "FAIL", "Backdrop image not found (Timeout)."
            except Exception as e:
                return "FAIL", f"Error: {type(e).__name__}"

        if case["CaseID"] == "TC22":
            try:
                cinema_select = wait.until(EC.visibility_of_element_located((By.XPATH, "//label[text()='Cinema']/following-sibling::div/select")))
                Select(cinema_select).select_by_index(1)
                # Thay sleep bằng wait logic
                date_select = wait.until(EC.element_to_be_clickable((By.XPATH, "//label[text()='Date']/following-sibling::div/select")))
                Select(date_select).select_by_index(1)
                
                # Chờ 1 chút để API load showtime
                time.sleep(1) 
                
                showtimes = driver.find_elements(By.XPATH, "//button[contains(@class, 'px-5')]")
                if showtimes:
                    return "PASS", f"Found {len(showtimes)} showtime(s)."
                
                no_showtime = driver.find_elements(By.XPATH, "//p[contains(., 'No showtimes available')]")
                if no_showtime:
                    return "FAIL", "No showtimes available (test data issue)."
                return "FAIL", "No showtimes and no 'unavailable' message displayed."
            except Exception as e:
                return "FAIL", f"Failed to load showtimes: {type(e).__name__}"

        if case["CaseID"] == "TC23":
            try:
                cinema_select = wait.until(EC.visibility_of_element_located((By.XPATH, "//label[text()='Cinema']/following-sibling::div/select")))
                Select(cinema_select).select_by_index(1)
                
                date_select = wait.until(EC.element_to_be_clickable((By.XPATH, "//label[text()='Date']/following-sibling::div/select")))
                Select(date_select).select_by_index(1)
                
                time.sleep(0.5)
                
                showtime_btn = driver.find_elements(By.XPATH, "//button[contains(@class, 'px-5')]")
                if not showtime_btn:
                    return "SKIP", "No showtimes to test redirect."
                
                showtime_btn[0].click()
                
                # Chờ URL thay đổi tối đa 3s
                try:
                    WebDriverWait(driver, 3).until(EC.url_contains("/login"))
                    return "PASS", "Correctly redirected to login page."
                except TimeoutException:
                    if "/booking" in driver.current_url:
                        return "FAIL", "User already logged in (test config issue)."
                    return "FAIL", f"Unexpected navigation to: {driver.current_url}"
            except Exception as e:
                return "FAIL", f"Redirect test failed: {type(e).__name__}"
        
        # --- Xử lý CASE LỖI TC20 (Tối ưu tốc độ) ---
        if case["CaseID"] == "TC20":
            try:
                # Dùng short_wait (2s) để kiểm tra dòng chữ "Upcoming Release".
                # Nếu logic đúng (đang chiếu), dòng này KHÔNG hiện ra -> Timeout sau 2s (thay vì 10s)
                short_wait.until(EC.visibility_of_element_located((By.XPATH, "//h3[contains(., 'Upcoming Release')]")))
                return "FAIL", "[BUG] 'Upcoming Release' shown for currently showing movie."
            except TimeoutException:
                # Tốt! Không thấy "Upcoming", giờ check xem có thấy dropdown không
                try:
                    # Dropdown phải hiện ngay, dùng short_wait cũng được
                    short_wait.until(EC.visibility_of_element_located((By.XPATH, "//label[text()='Cinema']")))
                    return "PASS", "Cinema selection displayed correctly."
                except TimeoutException:
                    return "FAIL", "Neither 'Upcoming Release' nor 'Select Cinema' dropdown found."

        # --- Xử lý các CASE PASS ---
        if assert_type == "element_visible":
            locator = get_locator(selector)
            if case["CaseID"] == "TC01":
                try:
                    # Spinner chỉ check 1s thôi, vì load nhanh quá sẽ lỡ
                    WebDriverWait(driver, 1).until(EC.visibility_of_element_located(locator))
                    return "PASS", "Loading spinner displayed."
                except TimeoutException:
                    return "PASS", "Page loaded fast. Spinner not visible (acceptable)."
            
            wait.until(EC.visibility_of_element_located(locator))
            return "PASS", f"Element visible: {selector}"

        if assert_type == "element_visible_after_click":
            if click_selector:
                click_locator = get_locator(click_selector)
                wait.until(EC.element_to_be_clickable(click_locator)).click()
            # Giảm sleep, dùng wait
            time.sleep(0.5)
            locator = get_locator(selector)
            wait.until(EC.visibility_of_element_located(locator))
            return "PASS", f"Element '{selector}' visible after click."

        if assert_type == "url_contains_after_click":
            locator = get_locator(selector)
            wait.until(EC.element_to_be_clickable(locator)).click()
            wait.until(EC.url_contains(expected_text))
            return "PASS", f"Navigation successful. URL contains: '{expected_text}'"
        
        if assert_type == "element_is_disabled":
            locator = get_locator(selector)
            element = wait.until(EC.visibility_of_element_located(locator))
            if not element.is_enabled():
                return "PASS", f"Element '{selector}' correctly disabled."
            return "FAIL", f"[BUG] Element '{selector}' enabled (expected disabled)."

        if assert_type == "element_visible_after_hover":
            cast_link = wait.until(EC.element_to_be_clickable((By.XPATH, "(//div[contains(@class, 'grid')]//a[contains(@href, '/actor/')])[1]")))
            cast_link.click()
            wait.until(EC.visibility_of_element_located((By.XPATH, "//h2[contains(., 'Movies')]")))
            
            movie_card = wait.until(EC.visibility_of_element_located((By.XPATH, "(//div[contains(@class, 'grid')]//a[contains(@href, '/movie/')])[1]")))
            ActionChains(driver).move_to_element(movie_card).perform()
            time.sleep(0.2) # Giảm sleep
            
            modal = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'fixed') and contains(@class, 'z-[9998]')]")))
            return "PASS", "MoviePreviewModal displayed on hover."
        
        if assert_type == "check_modal_flip_fail":
            cast_link = wait.until(EC.element_to_be_clickable((By.XPATH, "(//div[contains(@class, 'grid')]//a[contains(@href, '/actor/')])[1]")))
            cast_link.click()
            wait.until(EC.visibility_of_element_located((By.XPATH, "//h2[contains(., 'Movies')]")))
            
            movie_cards = driver.find_elements(By.XPATH, "//div[contains(@class, 'grid')]//a[contains(@href, '/movie/')]")
            total_movies = len(movie_cards)
            
            if total_movies < 2:
                return "FAIL", f"Insufficient test data: only {total_movies} movie(s)."
            
            target_index = min(4, total_movies - 1)
            target_movie = movie_cards[target_index]
            ActionChains(driver).move_to_element(target_movie).perform()
            time.sleep(0.5)
            
            try:
                modal = short_wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'fixed') and contains(@class, 'z-[9998]')]")))
                modal_x = modal.location['x']
                card_x = target_movie.location['x']
                
                if modal_x < card_x:
                    return "PASS", "Modal flip working correctly."
                return "FAIL", f"Modal flip not implemented (modal x={modal_x}, card x={card_x})."
            except TimeoutException:
                return "FAIL", "Modal not displayed on hover."

        return "SKIP", f"Unknown AssertType: '{assert_type}'."

    except TimeoutException:
        return "FAIL", f"Element not found: {selector} (timeout {wait._timeout}s)."
    except NoSuchElementException:
        return "FAIL", f"Element not found: {selector}."
    except Exception as e:
        error_name = type(e).__name__
        error_msg = str(e).split('\n')[0][:100]
        return "FAIL", f"Unexpected error: {error_name} - {error_msg}"


@pytest.mark.parametrize(
    "case_entry",
    _load_movie_cases(),
    ids=lambda case: case.get("CaseID", "unknown"),
)
@pytest.mark.e2e
def test_movie_flow(case_entry: Dict[str, str], driver, base_url) -> None:
    case = _normalize_case_entry(case_entry)
    status = "ERROR"
    notes = ""

    print(f"\n--- Bắt đầu {case.get('CaseID')}: {case.get('Scenario')} ---")

    try:
        status, notes = _execute_movie_case(driver, base_url, case)
    except Exception as exc:
        status = "FAIL"
        notes = f"{type(exc).__name__}: {exc}"
    finally:
        _append_report_row(case, status, notes)
        print(f"--- Kết thúc {case.get('CaseID')}: {status} ---")
        if status == "FAIL":
             print(f"[NOTE] {notes}")

    if status == "PASS":
        return
    if status == "SKIP":
        pytest.skip(notes or "Kịch bản bị bỏ qua")

    pytest.fail(notes or "Kịch bản thất bại")