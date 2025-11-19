from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- Test Data ---
# ID cho Phim Đang Chiếu (Dùng chung cho cả Upcoming nếu bạn muốn test giao diện)
MOVIE_NOW_SHOWING = "05705299-3003-464c-8dc3-d6d8ff7d9905"
MOVIE_UPCOMING = "05705299-3003-464c-8dc3-d6d8ff7d9905" 
MOVIE_INVALID = "invalid-uuid-123"

# --- Column Headers ---
HEADERS = [
    "CaseID",
    "Feature",
    "Scenario",
    "Preconditions",
    "Steps",
    "ExpectedResult",
    "Data_MovieID",        # ID phim để tải trang
    "Data_AssertSelector", # Selector CSS/XPath để kiểm tra
    "Data_ExpectedText",   # Văn bản mong đợi (cho assert text hoặc URL)
    "Data_ClickSelector",  # Selector cho phần tử cần click (nếu có)
    "AssertType",          # Loại kiểm tra (logic cho Selenium)
]

# --- 32 Test Cases Definition ---
CASES = [
    # I. Tải Trang & Giao diện Chung (6 cases)
    {
        "CaseID": "TC01",
        "Feature": "Loading",
        "Scenario": "Hiển thị spinner khi tải trang",
        "Preconditions": "Mạng chậm",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Biểu tượng spinner toàn trang hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//div[contains(@class, 'animate-spin')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC02",
        "Feature": "Loading",
        "Scenario": "Hiển thị lỗi khi ID phim không tồn tại",
        "Preconditions": "Movie ID không có trong DB",
        "Steps": f"Điều hướng đến /movie/{MOVIE_INVALID}",
        "ExpectedResult": "Thông báo 'Oops! Movie not found' hiển thị",
        "Data_MovieID": MOVIE_INVALID,
        "Data_AssertSelector": "//h2[contains(., 'Oops! Movie not found')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC03",
        "Feature": "Layout",
        "Scenario": "Breadcrumb hiển thị chính xác",
        "Preconditions": "Phim hợp lệ",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Breadcrumb hiển thị 'Home / [Tên Phim]'",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//nav[@aria-label='Breadcrumb']",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC04",
        "Feature": "Layout",
        "Scenario": "Hiển thị thanh bên 'Phim tương tự'",
        "Preconditions": "Phim hợp lệ",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Mục 'You Might Also Like' hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//h3[contains(., 'You Might Also Like')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC05",
        "Feature": "Layout",
        "Scenario": "Header (GlobalSearch) hiển thị",
        "Preconditions": "Phim hợp lệ",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Header (NewHeader) hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//header[contains(@class, 'sticky')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC06",
        "Feature": "Layout",
        "Scenario": "Footer hiển thị",
        "Preconditions": "Phim hợp lệ",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Footer (chứa 'CinesTech') hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//footer",
        "AssertType": "element_visible",
    },

    # II. MovieHero (Thông tin phim) (5 cases)
    {
        "CaseID": "TC07",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị thông tin phim chính xác (Tên, Mô tả, Poster)",
        "Preconditions": "Phim hợp lệ",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Tên phim, poster và mô tả hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//h1[contains(@class, 'text-4xl')]", 
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC08",
        "Feature": "MovieHero",
        "Scenario": "Mở modal trailer khi nhấn 'Watch Trailer'",
        "Preconditions": "Phim có trailer",
        "Steps": "Nhấn nút 'Watch Trailer'",
        "ExpectedResult": "Modal chứa iframe trailer hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//iframe[contains(@src, 'youtube')]",
        "Data_ClickSelector": "//button[contains(., 'Watch Trailer')]",
        "AssertType": "element_visible_after_click",
    },
    {
        "CaseID": "TC09",
        "Feature": "MovieHero",
        "Scenario": "[FAIL] Scroll lock not implemented when trailer modal opens",
        "Preconditions": "Movie has trailer available",
        "Steps": "1. Click 'Watch Trailer' button\n2. Verify modal opens with YouTube iframe\n3. Check document.body.style.overflow property",
        "ExpectedResult": "[DESIGNED TO FAIL] Body overflow should be 'hidden' to prevent background scrolling, but feature is not implemented",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_ClickSelector": "//button[contains(., 'Watch Trailer')]",
        "AssertType": "check_scroll_lock_fail",
    },
    {
        "CaseID": "TC10",
        "Feature": "MovieHero",
        "Scenario": "[FAIL] Backdrop image validation",
        "Preconditions": "Movie has backdrop image in database",
        "Steps": "1. Navigate to movie detail page\n2. Locate backdrop image in MovieHero section\n3. Verify image src attribute is valid URL",
        "ExpectedResult": "[DESIGNED TO FAIL] Backdrop image should load with valid src, but may fail due to CDN or image processing issues",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//div[contains(@class, 'absolute')]//img",
        "AssertType": "check_backdrop_fail",
    },
    {
        "CaseID": "TC11",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị đúng mác giới hạn độ tuổi",
        "Preconditions": "Phim hợp lệ",
        "Steps": "Kiểm tra mác tuổi trên poster",
        "ExpectedResult": "Mác tuổi (ví dụ 18+) hiển thị với màu đúng",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//div[contains(@class, 'bg-red-600')]", 
        "AssertType": "element_visible",
    },

    # III. CastSection & ActorDetail (6 cases)
    {
        "CaseID": "TC12",
        "Feature": "CastSection",
        "Scenario": "Hiển thị danh sách diễn viên",
        "Preconditions": "Phim hợp lệ",
        "Steps": "Cuộn xuống mục 'Cast & Crew'",
        "ExpectedResult": "Danh sách diễn viên (tên và ảnh) hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//h3[contains(., 'Cast & Crew')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC13",
        "Feature": "CastSection",
        "Scenario": "Điều hướng đến trang chi tiết diễn viên",
        "Preconditions": "Phim có diễn viên",
        "Steps": "Nhấn vào diễn viên đầu tiên trong danh sách",
        "ExpectedResult": "URL thay đổi thành '/actor/$actorId'",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "(//div[contains(@class, 'grid')]//a[contains(@href, '/actor/')])[1]",
        "Data_ExpectedText": "/actor/",
        "AssertType": "url_contains_after_click",
    },
    {
        "CaseID": "TC14",
        "Feature": "ActorDetail",
        "Scenario": "Trang chi tiết diễn viên hiển thị thông tin",
        "Preconditions": "Đã điều hướng từ TC13",
        "Steps": "Chờ trang ActorDetail tải",
        "ExpectedResult": "Tên diễn viên và tiểu sử (nếu có) hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//h1[contains(@class, 'text-2xl')]",
        "Data_ClickSelector": "(//div[contains(@class, 'grid')]//a[contains(@href, '/actor/')])[1]",
        "AssertType": "element_visible_after_click",
    },
    {
        "CaseID": "TC15",
        "Feature": "ActorDetail",
        "Scenario": "Trang chi tiết diễn viên hiển thị danh sách phim",
        "Preconditions": "Đã điều hướng từ TC13",
        "Steps": "Cuộn xuống mục 'Movies'",
        "ExpectedResult": "Danh sách phim của diễn viên hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//h2[contains(., 'Movies')]",
        "Data_ClickSelector": "(//div[contains(@class, 'grid')]//a[contains(@href, '/actor/')])[1]",
        "AssertType": "element_visible_after_click",
    },
    {
        "CaseID": "TC16",
        "Feature": "ActorDetail",
        "Scenario": "MoviePreviewModal hiển thị khi hover phim",
        "Preconditions": "Trang ActorDetail",
        "Steps": "Di chuột (hover) vào phim đầu tiên trong danh sách",
        "ExpectedResult": "Modal xem trước (MoviePreviewModal) hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//div[contains(@class, 'fixed') and contains(@class, 'z-[9998]')]",
        "AssertType": "element_visible_after_hover",
    },
    {
        "CaseID": "TC17",
        "Feature": "ActorDetail",
        "Scenario": "[FAIL] MoviePreviewModal lật vị trí khi ở gần lề phải",
        "Preconditions": "Trang ActorDetail",
        "Steps": "Hover vào phim ở cột cuối cùng bên phải",
        "ExpectedResult": "Modal hiển thị bên trái con trỏ (thay vì bên phải)",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "AssertType": "check_modal_flip_fail",
    },

    # IV. BookingSection (Đặt vé) (7 cases)
    {
        "CaseID": "TC18",
        "Feature": "BookingSection",
        "Scenario": "(Phim Sắp Chiếu) Hiển thị 'Upcoming Release'",
        "Preconditions": "Phim là phim sắp chiếu",
        "Steps": f"Điều hướng đến /movie/{MOVIE_UPCOMING}",
        "ExpectedResult": "Thông báo 'Upcoming Release' hiển thị",
        "Data_MovieID": MOVIE_UPCOMING,
        "Data_AssertSelector": "//h3[contains(., 'Upcoming Release')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC19",
        "Feature": "BookingSection",
        "Scenario": "(Phim Đang Chiếu) Hiển thị 'Select Cinema'",
        "Preconditions": "Phim là phim đang chiếu",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}",
        "ExpectedResult": "Dropdown 'Select Cinema' hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//label[text()='Cinema']/following-sibling::div/select",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC20",
        "Feature": "BookingSection",
        "Scenario": "[FAIL] Vẫn hiển thị 'Select Cinema' nếu hook `useMovieShowTimes` rỗng",
        "Preconditions": "Phim đang chiếu nhưng `useMovieShowTimes` (hook ban đầu) rỗng",
        "Steps": "Điều hướng đến phim (mô phỏng API trả về rỗng)",
        "ExpectedResult": "Dropdown 'Select Cinema' vẫn hiển thị (không bị báo 'Upcoming')",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//label[text()='Cinema']/following-sibling::div/select",
        "AssertType": "check_upcoming_logic_fail", 
    },
    {
        "CaseID": "TC21",
        "Feature": "BookingSection",
        "Scenario": "(Phim Đang Chiếu) Dropdown 'Date' bị vô hiệu hóa",
        "Preconditions": "Chưa chọn rạp (Cinema)",
        "Steps": "Kiểm tra dropdown 'Date'",
        "ExpectedResult": "Dropdown 'Date' có thuộc tính 'disabled'",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//label[text()='Date']/following-sibling::div/select",
        "AssertType": "element_is_disabled",
    },
    {
        "CaseID": "TC22",
        "Feature": "BookingSection",
        "Scenario": "[FAIL] Không có suất chiếu cho một số ngày",
        "Preconditions": "Phim đang chiếu nhưng không có suất chiếu",
        "Steps": "Chọn Cinema và Date",
        "ExpectedResult": "[DESIGNED TO FAIL] Có thể không có suất chiếu (hiển thị 'No showtimes available')",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//button[contains(@class, 'px-5')]",
        "AssertType": "element_visible_after_selection_fail",
    },
    {
        "CaseID": "TC23",
        "Feature": "BookingSection",
        "Scenario": "[FAIL] Chuyển hướng khi chưa login (cần implement)",
        "Preconditions": "Người dùng chưa đăng nhập",
        "Steps": "Chọn Cinema, Date, click showtime",
        "ExpectedResult": "[DESIGNED TO FAIL] Mong đợi chuyển đến /login nhưng có thể không được implement",
        "Data_MovieID": "05705299-3003-464c-8dc3-d6d8ff7d9905",
        "Data_ExpectedText": "/login",
        "AssertType": "url_redirect_fail",
    },

    # V. Kiểm tra chi tiết bổ sung (Success Cases - 8 cases mới)
    {
        "CaseID": "TC25",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị Đạo diễn (Director)",
        "Preconditions": "Phim có thông tin đạo diễn",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}; Kiểm tra thông tin Director",
        "ExpectedResult": "Tên đạo diễn được hiển thị (ví dụ sau nhãn 'Director:')",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//span[contains(text(), 'Director')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC26",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị Thể loại phim (Genres)",
        "Preconditions": "Phim có ít nhất 1 thể loại",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}; Kiểm tra danh sách thể loại",
        "ExpectedResult": "Các thẻ thể loại (Pills) hiển thị",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        # Tìm element có class rounded-full (thường dùng cho pill tag trong code của bạn)
        "Data_AssertSelector": "//div[contains(@class, 'flex')]//span[contains(@class, 'rounded-full')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC27",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị Điểm đánh giá (Rating Score)",
        "Preconditions": "Phim đã có đánh giá",
        "Steps": f"Điều hướng đến /movie/{MOVIE_NOW_SHOWING}; Kiểm tra điểm số",
        "ExpectedResult": "Điểm số dạng 'X / 10' hiển thị cùng ngôi sao",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//span[contains(text(), '/ 10')]",
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC28",
        "Feature": "Layout",
        "Scenario": "Điều hướng qua phim tương tự (Similar Movies)",
        "Preconditions": "Danh sách phim tương tự hiển thị",
        "Steps": "Click vào phim đầu tiên trong danh sách 'You Might Also Like'",
        "ExpectedResult": "URL thay đổi sang ID của phim mới",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "(//h3[contains(., 'You Might Also Like')]/following-sibling::div//a)[1]",
        "Data_ExpectedText": "/movie/",
        "AssertType": "url_contains_after_click",
    },
    {
        "CaseID": "TC29",
        "Feature": "Layout",
        "Scenario": "Breadcrumb điều hướng về trang chủ",
        "Preconditions": "Đang ở trang chi tiết phim",
        "Steps": "Click vào link 'Movies' (hoặc Home) trên Breadcrumb",
        "ExpectedResult": "Quay về trang danh sách phim (URL không còn chứa ID phim cũ)",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//nav[@aria-label='Breadcrumb']//a[contains(@href, '/')]", # Link đầu tiên thường là Home
        "Data_ExpectedText": "cinestech.me", # Hoặc chỉ kiểm tra URL gốc
        "AssertType": "url_contains_after_click",
    },
    {
        "CaseID": "TC30",
        "Feature": "Layout",
        "Scenario": "Thanh tìm kiếm (Global Search) hoạt động",
        "Preconditions": "Header hiển thị",
        "Steps": "Nhập text vào ô tìm kiếm trên header",
        "ExpectedResult": "Ô input nhận giá trị và không bị disable",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//input[@type='text' and @placeholder='Search movies...']",
        "AssertType": "element_visible", # Chỉ kiểm tra nó tồn tại và visible
    },
    {
        "CaseID": "TC31",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị Ngày phát hành (Release Date)",
        "Preconditions": "Phim có ngày phát hành",
        "Steps": "Kiểm tra thông tin ngày tháng (có icon lịch)",
        "ExpectedResult": "Ngày phát hành hiển thị (chứa icon Calendar hoặc text định dạng ngày)",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        # Tìm text nằm cạnh icon Calendar (thường SVG) hoặc text chứa năm 202X
        "Data_AssertSelector": "//div[contains(@class, 'flex')]//span[contains(text(), '20')]", 
        "AssertType": "element_visible",
    },
    {
        "CaseID": "TC32",
        "Feature": "MovieHero",
        "Scenario": "Hiển thị Thời lượng phim (Duration)",
        "Preconditions": "Phim có thời lượng",
        "Steps": "Kiểm tra thông tin thời lượng (có icon đồng hồ)",
        "ExpectedResult": "Thời lượng hiển thị (vd: '2h 15m')",
        "Data_MovieID": MOVIE_NOW_SHOWING,
        "Data_AssertSelector": "//span[contains(text(), 'm') and (contains(text(), 'h') or contains(text(), 'min'))]",
        "AssertType": "element_visible",
    },
]

# --- Hàm ghi file Excel (Giữ nguyên format) ---
OUTPUT_PATH = Path(__file__).with_name("movie_test_cases.xlsx")

def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MovieCases"

    # --- Định dạng tiêu đề ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 40

    # --- Thêm dữ liệu ---
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    
    for row_data in CASES:
        row = [row_data.get(column, "") for column in HEADERS]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Tô màu các case lỗi
            if "[FAIL]" in row_data.get("Scenario", ""):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")

    # --- Tự động điều chỉnh độ rộng cột ---
    column_widths = {
        'A': 8,  # CaseID
        'B': 15, # Feature
        'C': 40, # Scenario
        'D': 35, # Preconditions
        'E': 40, # Steps
        'F': 40, # ExpectedResult
        'G': 38, # Data_MovieID
        'H': 40, # Data_AssertSelector
        'I': 20, # Data_ExpectedText
        'J': 40, # Data_ClickSelector
        'K': 25, # AssertType
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(OUTPUT_PATH)
        print(f"Thành công: Đã tạo file 'movie_test_cases.xlsx' với {len(CASES)} test cases.")
    except PermissionError:
        print(f"LỖI: Không thể ghi file. Vui lòng đóng file 'movie_test_cases.xlsx' nếu nó đang mở.")

if __name__ == "__main__":
    main()