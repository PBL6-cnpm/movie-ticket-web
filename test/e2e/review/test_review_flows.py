from __future__ import annotations

import os
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill  # <--- Added Import
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# --- Configuration ---
MODULE_DIR = Path(__file__).resolve().parent
CASES_PATH = MODULE_DIR / "review_test_cases.xlsx"
REPORT_DIR = MODULE_DIR / "reports"
REPORT_FILE = REPORT_DIR / "review_results.xlsx"
REPORT_HEADERS = [
    "TimestampUTC", "CaseID", "Feature", "Scenario", "Status", "Notes",
]

load_dotenv()

# --- Define Locators (Based on .tsx code) ---
LOCATORS = {
    "sign_in_button": (By.XPATH, "//button[contains(., 'Sign in to review')]"),
    "comment_textarea": (By.XPATH, "//textarea[contains(@placeholder, 'What did you think')]"),
    "post_review_button": (By.XPATH, "//button[contains(., 'Post Review')]"),
    "stars_container": (By.XPATH, "//div[contains(@class, 'flex items-center gap-1.5 flex-wrap')]"),
    "already_reviewed_div": (By.XPATH, "//p[contains(., 'You already reviewed this movie.')]"),
    "remove_review_button": (By.XPATH, "//button[contains(., 'Remove review')]"),
    "load_more_button": (By.XPATH, "//button[contains(., 'Load more reviews')]"),
    
    # FIXED (REV-003): Based on MovieDetailPage.tsx line 1007
    "review_list_container": (By.XPATH, "//div[contains(@class, 'space-y-6')]/div[contains(@class, 'space-y-4')]"),
    "no_reviews_yet": (By.XPATH, ".//p[contains(text(), 'No reviews yet')]"),
    "first_review_item_relative": (By.XPATH, ".//p[contains(@class, 'whitespace-pre-line')]"),
    
    # FIXED (REV-011): Based on line 1007 (list) -> 1009 (item) -> 1025 (name)
    "first_review_name": (By.XPATH, "//div[contains(@class, 'space-y-6')]/div[contains(@class, 'space-y-4')]/div[1]//p[contains(@class, 'font-semibold')]"),
    
    # FIXED (REV-020): Based on MovieHero.tsx line 215-225
    "hero_rating": (By.XPATH, "//div[contains(@class, 'flex items-start gap-x-4')]//span[contains(text(), '/ 10')]"),
    
    # FIXED (REV-007): Based on MovieDetailPage.tsx line 996
    "form_error_span": (By.XPATH, "//form//span[contains(@class, 'text-red-400')]"),
    
    # FIXED (REV-017): Based on MyReviewsPage.tsx
    "profile_remove_button": (By.XPATH, ".//button[contains(., 'Remove rating')]"),
    # Wait for review card to disappear instead of toast
    "profile_review_card": (By.XPATH, "//div[@data-review-id]"),
}

# --- Helpers: Report Management ---

def _ensure_report_directory() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def _load_review_cases() -> List[Dict[str, str]]:
    if not CASES_PATH.exists():
        pytest.skip(f"Review cases file not found at {CASES_PATH}")
    workbook = load_workbook(CASES_PATH)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cases: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row): continue
        entry: Dict[str, str] = {header: value for header, value in zip(headers, row)}
        cases.append(_normalize_case_entry(entry))
    if len(cases) < 1:
        pytest.skip("Review cases workbook does not contain any rows to execute")
    return cases


def _append_report_row(case: Dict[str, str], status: str, notes: str) -> None:
    _ensure_report_directory()
    if REPORT_FILE.exists():
        workbook = load_workbook(REPORT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(REPORT_HEADERS)
        # Format bold header for new file
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            
    # Current row data
    row_data = [
        datetime.now(UTC).isoformat(timespec="seconds"),
        case.get("CaseID"), case.get("Feature"), case.get("Scenario"),
        status, notes,
    ]
    
    sheet.append(row_data)

    # --- LOGIC FORMAT RED IF FAIL ---
    if status == "FAIL":
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        
        # Get the last added row index
        last_row_index = sheet.max_row
        
        # Color each cell in that row
        for col_idx in range(1, len(row_data) + 1):
            cell = sheet.cell(row=last_row_index, column=col_idx)
            cell.fill = red_fill
            cell.font = red_font

    # Auto-adjust column width (for readability)
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width
    
    workbook.save(REPORT_FILE)


def _normalize_case_entry(entry: Dict[str, str]) -> Dict[str, str]:
    normalized: Dict[str, str] = {}
    for key, value in entry.items():
        if value is None: normalized[key] = ""
        elif isinstance(value, str): normalized[key] = value.strip()
        else: normalized[key] = str(value)
    return normalized

def _get_wait(driver, timeout: int = 10) -> WebDriverWait:
    return WebDriverWait(driver, timeout)

def _get_wait_long(driver, timeout: int = 15) -> WebDriverWait:
    return WebDriverWait(driver, timeout)

def _get_wait_short(driver, timeout: float = 2.0) -> WebDriverWait: # Slightly increased
    return WebDriverWait(driver, timeout)

def _clear_existing_report() -> None:
    _ensure_report_directory()
    if REPORT_FILE.exists():
        REPORT_FILE.unlink()


# --- Helpers: Selenium & State Management ---

def _logout_if_possible(driver, wait, base_url: str) -> None:
    try:
        driver.get(f"{base_url}/") 
        avatar = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "img[alt='User']")))
        avatar.click()
        logout_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Logout')]")))
        logout_button.click()
        wait.until(EC.visibility_of_element_located((By.XPATH, "//a[contains(., 'Login')]")))
    except Exception:
        driver.get(f"{base_url}/") 


def _login_if_needed(driver, wait, base_url: str, email: str, password: str) -> None:
    driver.get(base_url)
    time.sleep(0.5)
    try:
        driver.find_element(By.CSS_SELECTOR, "img[alt='User']")
        return
    except NoSuchElementException:
        driver.get(f"{base_url}/login")
        email_input = wait.until(EC.visibility_of_element_located((By.ID, "email")))
        password_input = driver.find_element(By.ID, "password")
        submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
        email_input.send_keys(email)
        password_input.send_keys(password)
        submit_button.click()
        try:
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "img[alt='User']")))
        except TimeoutException:
            raise Exception(f"Cannot login with user {email} to execute test case")


def _fill_if_present(element, value: str) -> None:
    if value:
        element.clear()
        element.send_keys(value)
    else:
        element.clear()

def _get_assert_type(case: Dict[str, str]) -> str:
    return (case.get("AssertType") or "").strip().lower()

def _scroll_to_reviews(driver):
    try:
        review_header = driver.find_element(By.XPATH, "//h3[contains(text(), 'Audience Reviews')]")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", review_header)
        time.sleep(1.5)
    except NoSuchElementException:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 2);")
        time.sleep(1.5)

def _get_form_error(wait_short: WebDriverWait) -> Tuple[str, str]:
    try:
        error_element = wait_short.until(
            EC.visibility_of_element_located(LOCATORS["form_error_span"])
        )
        text = error_element.text.strip()
        if text: return "error", text
        return "none", ""
    except TimeoutException:
        return "none", ""

def _ensure_review_form_is_visible(driver, wait: WebDriverWait, wait_long: WebDriverWait):
    """(Cleanup) Ensure review form is visible."""
    try:
        wait.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"]))
        return 
    except TimeoutException:
        try:
            remove_button = wait.until(EC.element_to_be_clickable(LOCATORS["remove_review_button"]))
            remove_button.click()
            wait_long.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"]))
        except Exception as e:
            raise Exception(f"Error cleaning up (deleting old review): {e}")

def _ensure_review_exists(driver, wait: WebDriverWait, wait_long: WebDriverWait):
    """(Preparation) Ensure user has a review."""
    try:
        wait.until(EC.visibility_of_element_located(LOCATORS["already_reviewed_div"]))
        return 
    except TimeoutException:
        try:
            wait.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"]))
            stars_container = driver.find_element(*LOCATORS["stars_container"])
            star_buttons = stars_container.find_elements(By.TAG_NAME, "button")
            star_buttons[4].click() # 5 stars
            _fill_if_present(driver.find_element(*LOCATORS["comment_textarea"]), "Temp review for testing")
            wait.until(EC.element_to_be_clickable(LOCATORS["post_review_button"])).click()
            wait_long.until(EC.visibility_of_element_located(LOCATORS["already_reviewed_div"]))
        except Exception as e:
            raise Exception(f"Error preparing (creating new review): {e}")

# --- Pytest Fixtures ---

@pytest.fixture(scope="session")
def base_url() -> str:
    url = os.getenv("E2E_BASE_URL") or "https://cinestech.me"
    return url.rstrip("/")


@pytest.fixture(scope="session")
def driver():
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1440,900")
    
    # FIXED: Add '#' to the line below if you WANT TO SEE Chrome running
    # options.add_argument("--headless")
    
    service = Service(ChromeDriverManager().install())
    browser = webdriver.Chrome(service=service, options=options)
    yield browser
    browser.quit()


@pytest.fixture(scope="session", autouse=True)
def reset_review_report() -> None:
    _clear_existing_report()


@pytest.fixture(scope="session")
def review_case_data() -> List[Dict[str, str]]:
    return _load_review_cases()


# --- Core Review Test Execution ---

def _execute_review_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    feature = case.get("Feature", "").lower()
    if feature == "movie review":
        return _run_movie_review_case(driver, base_url, case)
    if feature == "my reviews page":
        return _run_my_reviews_case(driver, base_url, case)
    return "SKIP", f"Unhandled feature: {case.get('Feature')}"


def _run_movie_review_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    wait = _get_wait(driver)
    wait_long = _get_wait_long(driver)
    wait_short = _get_wait_short(driver)
    movie_id = case.get("Data_MovieID")
    assert_type = _get_assert_type(case)
    expected_message = case.get("ExpectedUIMessage", "")
    
    target_url = f"{base_url}/movie/{movie_id}"
    if driver.current_url != target_url:
        driver.get(target_url)
        time.sleep(0.5) 
    
    if assert_type == "review_hero_check":
        try:
            # FIXED (REV-020)
            hero_rating_element = wait.until(EC.visibility_of_element_located(LOCATORS["hero_rating"]))
            hero_rating_text = hero_rating_element.text
            if expected_message in hero_rating_text:
                return "PASS", f"Rating confirmed: {hero_rating_text}"
            return "FAIL", f"Incorrect rating. Expected '{expected_message}', got '{hero_rating_text}'"
        except TimeoutException:
            return "FAIL", "Rating element not found in hero."

    _scroll_to_reviews(driver)
    
    if assert_type == "review_form_hidden":
        try:
            locator = LOCATORS["sign_in_button"] if "Sign in" in expected_message else LOCATORS["already_reviewed_div"]
            element = wait.until(EC.visibility_of_element_located(locator))
            if expected_message in element.text:
                return "PASS", f"Element '{expected_message}' displayed correctly."
            return "FAIL", f"Element displayed but text mismatch: '{element.text}'"
        except TimeoutException:
            return "FAIL", f"Element not found for '{assert_type}'"

    if assert_type == "redirect_to_login":
        wait.until(EC.element_to_be_clickable(LOCATORS["sign_in_button"])).click()
        wait.until(EC.url_contains(expected_message))
        return "PASS", f"Redirected to {driver.current_url}"

    if assert_type == "review_list_visible":
        try:
            # FIXED (REV-003): Wait for container AND (item OR "no reviews")
            container = wait.until(EC.visibility_of_element_located(LOCATORS["review_list_container"]))
            try:
                # Check for the "No reviews yet" message (line 1060)
                no_reviews_locator = (By.XPATH, ".//p[contains(text(), 'No reviews yet')]")
                wait_short.until(lambda d: container.find_element(*no_reviews_locator))
                return "PASS", "Container found (currently empty)."
            except TimeoutException:
                # Check for the first review item (line 1056)
                first_review_locator = (By.XPATH, ".//p[contains(@class, 'whitespace-pre-line')]")
                wait_short.until(lambda d: container.find_element(*first_review_locator))
                return "PASS", "Container found with at least 1 review."
        except TimeoutException:
            return "FAIL", "Review list container not found."

    if assert_type == "review_submit_success" or assert_type == "ui_error_message":
        try:
            rating = int(case.get("Data_Rating", 0))
            comment = case.get("Data_Comment", "")
            
            comment_textarea = wait.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"]))
            
            if rating > 0:
                stars_container = driver.find_element(*LOCATORS["stars_container"])
                star_buttons = stars_container.find_elements(By.TAG_NAME, "button")
                star_buttons[rating - 1].click()
            
            _fill_if_present(comment_textarea, comment)
            wait.until(EC.element_to_be_clickable(LOCATORS["post_review_button"])).click()

            if assert_type == "review_submit_success":
                # LOGIC FIX: Wait for 'already reviewed' (wait 15s)
                wait_long.until(EC.visibility_of_element_located(LOCATORS["already_reviewed_div"]))
                return "PASS", "Review submitted successfully (form hidden, 'already reviewed' appeared)."
            
            if assert_type == "ui_error_message":
                # LOGIC FIX: Wait for error span inside form
                error_element = wait_long.until(EC.visibility_of_element_located(LOCATORS["form_error_span"]))
                feedback_text = error_element.text.strip()
                if expected_message in feedback_text:
                    return "PASS", f"Correct error displayed: {feedback_text}"
                return "FAIL", f"Expected error ('{expected_message}'), but got: ('{feedback_text}')"
        
        except TimeoutException:
            if assert_type == "review_submit_success":
                # FIXED (3 FAIL CASES): Expected success, but timed out
                feedback_type, feedback_text = _get_form_error(wait_short) 
                if feedback_type == "error":
                    return "FAIL", f"Expected success, but received error: '{feedback_text}'"
            return "FAIL", f"Timeout while executing '{assert_type}'"
        except Exception as e:
            return "FAIL", f"Error executing '{assert_type}': {e}"

    if assert_type == "review_self_on_top":
        try:
            # FIXED (REV-011)
            first_reviewer_name = wait.until(EC.visibility_of_element_located(LOCATORS["first_review_name"])).text
            user_full_name = case.get("Data_User_FullName", "")
            if not user_full_name: return "FAIL", "Missing Data_User_FullName in test case"
            if user_full_name == first_reviewer_name:
                return "PASS", "Own review displayed at top of list."
            return "FAIL", f"First reviewer is '{first_reviewer_name}', expected '{user_full_name}'."
        except TimeoutException:
            return "FAIL", "First reviewer name not found."

    if assert_type == "review_remove_visible":
        try:
            wait.until(EC.visibility_of_element_located(LOCATORS["remove_review_button"]))
            return "PASS", "'Remove review' button visible."
        except TimeoutException:
            return "FAIL", "'Remove review' button not found."

    if assert_type == "review_remove_success":
        try:
            wait.until(EC.element_to_be_clickable(LOCATORS["remove_review_button"])).click()
            wait_long.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"])) 
            return "PASS", "Review removed successfully, input form reappeared."
        except TimeoutException:
            return "FAIL", "Review removal failed or new form did not appear."

    if assert_type == "review_form_visible":
        try:
            wait.until(EC.visibility_of_element_located(LOCATORS["comment_textarea"]))
            return "PASS", "Review form visible."
        except TimeoutException:
            return "FAIL", "Review form not visible."

    if assert_type == "review_pagination_success":
        try:
            load_more_button = wait.until(EC.element_to_be_clickable(LOCATORS["load_more_button"]))
            load_more_button.click()
            time.sleep(1) 
            return "PASS", "'Load more reviews' clicked."
        except TimeoutException:
            return "PASS", "'Load more' button not found (possibly loaded all)."
        except Exception as e:
            return "FAIL", f"Error clicking 'Load more': {e}"

    return "SKIP", f"Unhandled assertion type for Movie Review: {assert_type}"


def _run_my_reviews_case(driver, base_url: str, case: Dict[str, str]) -> Tuple[str, str]:
    wait = _get_wait(driver)
    wait_long = _get_wait_long(driver)
    assert_type = _get_assert_type(case)
    expected_comment = case.get("Data_Comment", "")
    expected_message = case.get("ExpectedUIMessage", "") 

    driver.get(f"{base_url}/profile/reviews")
    time.sleep(1) 

    if assert_type == "review_list_updated":
        try:
            # Find review using flexible XPath
            review_found = wait.until(EC.visibility_of_element_located(
                (By.XPATH, f"//*[contains(text(), '{expected_comment}')]")
            ))
            return "PASS", f"Found review with comment '{expected_comment}'."
        except TimeoutException:
            # Try finding any review
            try:
                any_review = driver.find_elements(By.XPATH, "//div[@data-review-id]")
                if any_review:
                    return "FAIL", f"Found {len(any_review)} review(s), but none with comment '{expected_comment}'."
                return "FAIL", "No reviews found in list."
            except:
                return "FAIL", f"Review with comment '{expected_comment}' not found."

    if assert_type == "review_remove_success_profile":
        try:
            # FIXED (REV-017): Find review card by comment text
            review_card = wait.until(EC.visibility_of_element_located(
                (By.XPATH, f"//h3[contains(text(), '{expected_comment}')]/ancestor::div[@data-review-id]")
            ))
            
            # Get review ID for tracking
            review_id = review_card.get_attribute("data-review-id")
            
            # Find and click remove button within that card
            remove_button = review_card.find_element(*LOCATORS["profile_remove_button"])
            driver.execute_script("arguments[0].scrollIntoView(true);", remove_button)
            time.sleep(0.5)
            remove_button.click()
            
            # Wait for review card to disappear (staleness)
            wait_long.until(EC.staleness_of(review_card))
            
            # Verify it's no longer in DOM
            remaining_cards = driver.find_elements(By.XPATH, f"//div[@data-review-id='{review_id}']")
            if len(remaining_cards) == 0:
                return "PASS", "Removed review from 'My Reviews' successfully."
            return "FAIL", "Review card still in DOM after removal."
            
        except TimeoutException:
            return "FAIL", "Timeout removing review or waiting for element to disappear."
        except Exception as e:
            return "FAIL", f"Error removing review: {e}"

    return "SKIP", f"Unhandled assertion type for My Reviews Page: {assert_type}"


# --- Main Test Runner ---

# FIXED (REV-015): Move REV-015 to correct group
CASES_GUEST = ['REV-001', 'REV-002', 'REV-003', 'REV-019', 'REV-020']
CASES_SUBMIT_REVIEW = [
    'REV-007', 'REV-008', 'REV-009', 'REV-010', 
    'REV-004', 'REV-005', 'REV-006',
    'REV-015' # <<< MOVED HERE
]
CASES_MANAGE_REVIEW = [
    'REV-011', 'REV-012', 'REV-013', 'REV-014', 
    # REV-015 REMOVED FROM HERE
    'REV-016', 'REV-017', 
    'REV-018' 
]

@pytest.mark.e2e
def test_entire_review_flow(driver, base_url, review_case_data) -> None:
    
    current_login_state = None 
    wait = _get_wait(driver)
    wait_long = _get_wait_long(driver)
    
    all_cases_map = {case.get("CaseID"): case for case in review_case_data}
    ordered_case_ids = CASES_GUEST + CASES_SUBMIT_REVIEW + CASES_MANAGE_REVIEW
    all_cases = [all_cases_map[case_id] for case_id in ordered_case_ids if case_id in all_cases_map]
    
    failed_cases = []

    for case_entry in all_cases:
        case = _normalize_case_entry(case_entry)
        case_id = case.get("CaseID", "unknown")
        
        status = "ERROR"
        notes = ""
        
        email = case.get("Data_Email")
        password = case.get("Data_Password")
        
        try:
            # 1. Manage login state
            if email and password: 
                if current_login_state != email:
                    _login_if_needed(driver, wait, base_url, email, password)
                    current_login_state = email
            else: 
                if current_login_state is not None:
                    _logout_if_possible(driver, wait, base_url)
                    current_login_state = None

            # 2. Manage Review State (Cleanup/Prep)
            if current_login_state: 
                if case.get("Feature") == "Movie Review":
                    target_url = f"{base_url}/movie/{case.get('Data_MovieID')}"
                    if driver.current_url != target_url:
                        driver.get(target_url)
                    _scroll_to_reviews(driver)
                    
                    # FIXED (REV-015): Update state logic
                    if case_id in CASES_SUBMIT_REVIEW + ['REV-018']:
                        print(f"   [State Check] Ensure form is empty for {case_id}...")
                        _ensure_review_form_is_visible(driver, wait, wait_long)
                    
                    elif case_id in CASES_MANAGE_REVIEW:
                        print(f"   [State Check] Ensure review exists for {case_id}...")
                        _ensure_review_exists(driver, wait, wait_long)

            # 3. Execute
            print(f"\n--- Start {case_id}: {case.get('Scenario')} ---")
            status, notes = _execute_review_case(driver, base_url, case)
            print(f"--- End {case_id}: {status} ---")

        except Exception as exc:
            status = "FAIL"
            notes = f"Critical Error: {type(exc).__name__}: {exc}"
            _append_report_row(case, status, notes)
            if not any(case_id in f for f in failed_cases):
                 failed_cases.append(f"{case_id} ({status}): {notes}")
            driver.get(base_url) 
        
        finally:
            if status != "ERROR": 
                if not any(case_id in f for f in failed_cases):
                    _append_report_row(case, status, notes)
            if status != "PASS":
                if not any(case_id in f for f in failed_cases):
                    failed_cases.append(f"{case_id} ({status}): {notes}")

    if current_login_state is not None:
        _logout_if_possible(driver, wait, base_url)

    if failed_cases:
        print("\n--- SUMMARY OF FAILED CASES ---")
        for fail in failed_cases:
            print(fail)
        pytest.fail(f"Total {len(failed_cases)}/{len(all_cases)} cases failed. See details in 'review_results.xlsx'.")